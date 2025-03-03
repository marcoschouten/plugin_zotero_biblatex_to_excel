import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
import os


# Function to generate Excel file
def generate_excel(content):
    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define a list of colors for the entries
    colors = ["ADD8E6", "E0FFFF", "98FB98", "FFB6C1", "FFD700", "DDA0DD", "B0E0E6", "FFDEAD", "AFEEEE", "F0E68C"]

    # Define a border style with left and right borders
    side_border = Border(left=Side(style="thin"), right=Side(style="thin"))

    # Define a border style with left, right, and bottom borders
    bottom_border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))

    # Define a centered alignment
    centered_alignment = Alignment(horizontal="center")

    # Split the content into lines and process each line
    lines = content.strip().split("\n")
    start_row = 1
    current_col_index = 1  # Start in column A
    entries = []  # List to hold entries under each header

    # Process each line in the content
    for line in lines:
        if line.strip():  # Ensure the line is not empty
            # Check if the line is a header (all uppercase)
            if line.isupper():
                # If there are entries collected, add them as a string below the previous header
                if entries:
                    # Store the previous entries in the current column with a single color
                    for idx, entry in enumerate(entries, start=start_row + 1):
                        col_letter = openpyxl.utils.get_column_letter(current_col_index)
                        sheet[f"{col_letter}{idx}"] = entry

                        # Apply a single color and side border to all entries in the column
                        color = colors[(current_col_index - 1) % len(colors)]  # Use the same color for the entire column
                        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        sheet[f"{col_letter}{idx}"].fill = fill
                        sheet[f"{col_letter}{idx}"].border = side_border

                        # Apply a bottom border to the last entry in the column
                        if idx == start_row + len(entries):
                            sheet[f"{col_letter}{idx}"].border = bottom_border
                    entries = []  # Reset entries for the next header

                # Move to the next column for a new header
                current_col_index += 1
                # Set the header for the column
                col_letter = openpyxl.utils.get_column_letter(current_col_index)

                # Apply the same color, side border, and centered alignment to the header as the entries
                color = colors[(current_col_index - 1) % len(colors)]  # Use the same color for the entire column
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                sheet[f"{col_letter}{start_row}"].fill = fill
                sheet[f"{col_letter}{start_row}"].border = side_border
                sheet[f"{col_letter}{start_row}"].alignment = centered_alignment

                sheet[f"{col_letter}{start_row}"] = line
            else:
                # Add the entry to the list for the current concept
                entries.append(line)

    # Add any remaining entries for the last header
    if entries:
        for idx, entry in enumerate(entries, start=start_row + 1):
            col_letter = openpyxl.utils.get_column_letter(current_col_index)
            sheet[f"{col_letter}{idx}"] = entry

            # Apply a single color and side border to all entries in the column
            color = colors[(current_col_index - 1) % len(colors)]  # Use the same color for the entire column
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sheet[f"{col_letter}{idx}"].fill = fill
            sheet[f"{col_letter}{idx}"].border = side_border

            # Apply a bottom border to the last entry in the column
            if idx == start_row + len(entries):
                sheet[f"{col_letter}{idx}"].border = bottom_border

    # Save the workbook
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads", "output.xlsx")
    workbook.save(downloads_path)
    print(f"Excel file saved to {downloads_path}")


# Example usage
content = """
CONCEPT_CREATION_BLENDING
liuCompositionalVisual2023 | http://arxiv.org/abs/2206.01714
kumariMultiConceptCustomization2023a | https://ieeexplore.ieee.org/document/10203856/
dengFireFlowFast2024 | http://arxiv.org/abs/2412.07517
zhouFreeBlendAdvancing2025a | http://arxiv.org/abs/2502.05606

TITLE
smithNewConcept2026 | http://arxiv.org/abs/2602.07517
johnsonAnotherIdea2027 | http://arxiv.org/abs/2703.01689
"""

generate_excel(content)
