import bibtexparser
import os
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QTextEdit, QPushButton
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QClipboard
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment

# If a file output.txt exists, delete all its content
if os.path.exists("output.txt"):
    with open("output.txt", "w", encoding="utf-8") as output_file:
        pass

# Get all .bib files in the current folder
bib_files = [f for f in os.listdir(".") if f.endswith(".bib")]


# Function to process .bib files
def process_bib_files(bib_files):
    """
    Process each .bib file in the list by replacing '@online' with '@misc',
    sorting entries by the 'date' field, and writing the results to 'output.txt'.

    Parameters:
    bib_files (list): List of .bib file names to process.
    """
    for bib_file_name in bib_files:
        with open(bib_file_name, "r", encoding="utf-8") as bib_file:
            content = bib_file.read()
        content = content.replace("@online", "@misc")
        with open(bib_file_name, "w", encoding="utf-8") as bib_file:
            bib_file.write(content)

    for bib_file_name in bib_files:
        # Load the .bib file
        with open(bib_file_name, "r", encoding="utf-8") as bib_file:
            bib_database = bibtexparser.load(bib_file)

        # Sort entries by year extracted from the date field
        sorted_entries = sorted(bib_database.entries, key=lambda x: x.get("date", "9999-12-31"))
        bib_database.entries = sorted_entries
        # Extract just the filename from the full path
        filename_only = os.path.basename(bib_file_name)
        with open("output.txt", "a", encoding="utf-8") as output_file:
            output_file.write(f"{filename_only[:-4].upper()}\n")  # Add filename in caps without .bib
            for entry in bib_database.entries:
                entry_id = entry.get("ID", "No ID")
                entry_url = entry.get("url", "No URL")
                output_file.write(f"{entry_id} | {entry_url}\n")
            output_file.write("\n")  # Add a newline between each bib_file

    # Copy the content of the output.txt to the clipboard and return the content string
    return open("output.txt", "r", encoding="utf-8").read()


class DragDropWidget(QFrame):
    """
    A PyQt5 widget that allows drag-and-drop functionality for .bib files.
    """

    def __init__(self):
        super().__init__()

        self.setAcceptDrops(True)
        self.setWindowTitle("Drag-and-Drop .bib Files")
        self.setGeometry(100, 100, 600, 400)

        self.label = QLabel("Drag and drop your .bib files here", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.text_edit = QTextEdit(self)
        self.text_edit.setReadOnly(True)

        self.quit_button = QPushButton("Quit", self)
        self.quit_button.setFixedSize(150, 50)
        self.quit_button.clicked.connect(QApplication.instance().quit)

        self.copy_button = QPushButton("Copy to Clipboard", self)
        self.copy_button.setFixedSize(150, 50)
        self.copy_button.clicked.connect(self.copy_to_clipboard)

        self.save_button = QPushButton("Save to Excel", self)
        self.save_button.setFixedSize(150, 50)
        self.save_button.clicked.connect(self.save_to_excel)

        layout = QVBoxLayout(self)
        layout.addWidget(self.label)
        layout.addWidget(self.text_edit)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.copy_button)
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.quit_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [url.toLocalFile() for url in event.mimeData().urls() if url.toLocalFile().endswith(".bib")]
        process_bib_files(files)
        self.label.setText("Files processed!")

        # Read the output.txt and display it in the QTextEdit
        with open("output.txt", "r", encoding="utf-8") as output_file:
            self.text_edit.setPlainText(output_file.read())

    def copy_to_clipboard(self):
        clipboard = QApplication.clipboard()
        clipboard.setText(self.text_edit.toPlainText())

    def save_to_excel(self):
        """
        Read the content from output.txt and generate an Excel file.
        """
        with open("output.txt", "r", encoding="utf-8") as output_file:
            content = output_file.read()
        generate_excel(content)


# Function to generate Excel file
def generate_excel(content):
    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define a list of colors for the entries
    colors = ["ADD8E6", "E0FFFF", "98FB98", "FFB6C1", "FFD700", "DDA0DD", "B0E0E6", "FFDEAD", "AFEEEE", "F0E68C"]

    # Define a border style with left and right borders
    side_border = Border(left=Side(style="thin"), right=Side(style="thin"))

    # Define a border style with left, right, and bottom borders
    bottom_border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))

    # Define a centered alignment
    centered_alignment = Alignment(horizontal="center")

    # Split the content into lines and process each line
    lines = content.strip().split("\n")
    start_row = 1
    current_col_index = 1  # Start in column A
    entries = []  # List to hold entries under each header

    # Process each line in the content
    for line in lines:
        if line.strip():  # Ensure the line is not empty
            # Check if the line is a header (all uppercase)
            if line.isupper():
                # If there are entries collected, add them as a string below the previous header
                if entries:
                    # Store the previous entries in the current column with a single color
                    for idx, entry in enumerate(entries, start=start_row + 1):
                        col_letter = openpyxl.utils.get_column_letter(current_col_index)
                        sheet[f"{col_letter}{idx}"] = entry

                        # Apply a single color and side border to all entries in the column
                        color = colors[(current_col_index - 1) % len(colors)]  # Use the same color for the entire column
                        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        sheet[f"{col_letter}{idx}"].fill = fill
                        sheet[f"{col_letter}{idx}"].border = side_border

                        # Apply a bottom border to the last entry in the column
                        if idx == start_row + len(entries):
                            sheet[f"{col_letter}{idx}"].border = bottom_border
                    entries = []  # Reset entries for the next header

                # Move to the next column for a new header
                current_col_index += 1
                # Set the header for the column
                col_letter = openpyxl.utils.get_column_letter(current_col_index)

                # Apply the same color, side border, and centered alignment to the header as the entries
                color = colors[(current_col_index - 1) % len(colors)]  # Use the same color for the entire column
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                sheet[f"{col_letter}{start_row}"].fill = fill
                sheet[f"{col_letter}{start_row}"].border = side_border
                sheet[f"{col_letter}{start_row}"].alignment = centered_alignment

                sheet[f"{col_letter}{start_row}"] = line
            else:
                # Add the entry to the list for the current concept
                entries.append(line)

    # Add any remaining entries for the last header
    if entries:
        for idx, entry in enumerate(entries, start=start_row + 1):
            col_letter = openpyxl.utils.get_column_letter(current_col_index)
            sheet[f"{col_letter}{idx}"] = entry

            # Apply a single color and side border to all entries in the column
            color = colors[(current_col_index - 1) % len(colors)]  # Use the same color for the entire column
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sheet[f"{col_letter}{idx}"].fill = fill
            sheet[f"{col_letter}{idx}"].border = side_border

            # Apply a bottom border to the last entry in the column
            if idx == start_row + len(entries):
                sheet[f"{col_letter}{idx}"].border = bottom_border

    # Save the workbook
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads", "output_app.xlsx")
    workbook.save(downloads_path)
    print(f"Excel file saved to {downloads_path}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = DragDropWidget()
    widget.show()

    # Process the BibTeX files to update output.txt
    content = process_bib_files(bib_files)

    # Call the function to generate the Excel file
    generate_excel(content)

    sys.exit(app.exec_())
